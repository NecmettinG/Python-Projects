import time
import datetime
import pandas as pd
from openpyxl.workbook import Workbook
from openpyxl import load_workbook

input("press ENTER to start program")
time.sleep(15)
first_time = time.time()

#Open your excel file:
writer = load_workbook("C:/Users/gedik/Desktop/excel sorter/Dummy Database.xlsx") #Enter database's directory path here.
#Activate the first page:
page = writer.active

#PROCESS TIME SECTION:

#Read all data from Excel file with using pandas:
data = pd.read_excel("Dummy Database.xlsx", sheet_name="Sayfa1") #Enter your sheet name on sheet_name.

#Store all datas into a list. These datas will come from Finish_Time column.
data = data["Finish_Time"].tolist()

#Write the first time data to first row under Finish_Time key:
page.cell(row=2,column=13).value=data[0]

#Write all time datas to all cells under Finish_Time column:
print("1st loop started")
for x in range(3,len(data)+2):

    page.cell(row=x,column=13).value=data[x-3]
print("1st loop finished")

#DRAWING DATA FROM EXCEL FILE:
#to store all rows, create an empty list:
all_rows_list = []
#to store all personnel ids:
personnel_id = []
#to store all keys to each column:
keys_list = []

max_row_number = page.max_row #max_row_number of database
max_column_number = page.max_column #max_column_number of database

#some personnels have id "Y". This list will be used for storing personnel with id "Y":
list_with_y = []

print("2nd Loop started")
second_loop = 1 #for checking if loop is working or not. It also shows which row we are reading:
for x in range(2,max_row_number+1):
    print(second_loop)
    cells = page[x] #Store each row to cells as a tuple.

    #checking if personnel id is "Y" or not:
    if cells[8].value!="Y":
        temp_list = []
        loop_number=0 #It is for column indexes. But we didn't use range function because we have to reach values from inside of the cells tuple.
        for index in cells:
            if loop_number==8:#8 means personnel id index. When it reaches to 8, it means we are on personnel id index right now.
                #Personnels have "S" letter before their actual id except "Y" personnels. We have to get rid of "S" letter.
                temp_list.append(index.value.replace("S",""))
            else:
                temp_list.append(index.value)
            loop_number+=1
        #Store all personnel with ids except "Y"
        all_rows_list.append(temp_list)
        print(temp_list)

    #If personnel id is "Y":
    elif cells[8].value=="Y":
        print("Y block started"+"---"*50)
        temp_with_y = []
        for index in cells:
            temp_with_y.append(index.value)
        list_with_y.append(temp_with_y)
        print(temp_with_y)
        second_loop+=1
        continue

    #Store all personnels ids except "Y" without "S" letter:
    if cells[8].value.replace("S","") not in personnel_id and cells[8].value!="Y":
        print("new personnel id has been found")
        personnel_id.append(cells[8].value.replace("S",""))
        print(personnel_id)
    second_loop+=1

print("2nd loop finished")
#sort all rows list according to personnel id:
all_rows_list.sort(key = lambda x: x[8])

#extend the all rows list with y list.
all_rows_list.extend(list_with_y)

#Insert value row into index 0:
for key in page["1"]:
    keys_list.append(key.value)

all_rows_list.insert(0,keys_list)

#sorting personnel ids:
personnel_id.sort()

#Some personnels have Id "Y". Append it to last index:
personnel_id.append("Y")

#numbering each personnel accoring to id sorting:
print("3rd loop started")
for index in personnel_id:
    number = 1

    for x in all_rows_list:

        if x[8] == index:
            x[1] = number
            number+=1
print("3rd loop finished")

#overwrite all sorted rows to excel sheet:
print("4th loop started")
for rows in range(0,max_row_number):
    for columns in range(0,max_column_number):
        page.cell(row=rows+1,column=columns+1).value = all_rows_list[rows][columns]
print("4th loop finished")

second_time = time.time()
print(f"The runtime = {second_time-first_time}")

writer.save("Dummy Database.xlsx")

time.sleep(15)




